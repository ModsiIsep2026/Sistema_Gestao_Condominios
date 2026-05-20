import io
from datetime import datetime, date
from typing import Optional
import openpyxl # Esta biblioteca permite criar e editar ficheiros Excel
from fastapi import HTTPException
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas # Esta biblioteca permite criar e editar ficheiros PDF
from sqlalchemy.orm import Session

from app.models.categoria_despesa import CategoriaDespesa
from app.models.despesa import Despesa
from app.models.edificio import Edificio
from app.models.fornecedor import Fornecedor
from app.models.perfil import Perfil
from app.models.utilizador import Utilizador

PDF_LIMITE = 500
COR_CABECALHO = "1A2B6B"


def validar_gestor(db: Session, gestor_id: int) -> Utilizador:
    gestor = (db.query(Utilizador).join(Perfil, Utilizador.perfil_id == Perfil.id_perfil)
        
        .filter(
            Utilizador.id_utilizador == gestor_id,
            Utilizador.status == 1,
            Perfil.nome == "gestor",
            Perfil.status == 1,
        )
        .first()
    )

    if not gestor:
        raise HTTPException(403, "Apenas gestores podem criar relatorios")

    return gestor


def obter_despesas(db: Session,gestor_id: int,data_inicio: Optional[date],data_fim: Optional[date],limit: Optional[int] = None,):
    query = db.query(Despesa).filter(Despesa.status == 1,Despesa.gestor_id == gestor_id,)


    # Filtrar data
    if data_inicio:
        query = query.filter(Despesa.data_despesa >= data_inicio)
    if data_fim:
        query = query.filter(Despesa.data_despesa <= data_fim)

    query = query.order_by(Despesa.data_despesa.desc())

    if limit is not None:
        query = query.limit(limit) # Definir limites de quantidade de dados exportados

    return query.all()


def formatar_dados_export(db: Session, despesa: Despesa) -> dict:
    edificio = db.query(Edificio).filter(Edificio.id_edificio == despesa.edificio_id).first()
    categoria = db.query(CategoriaDespesa).filter(CategoriaDespesa.id_categoria == despesa.categoria_id).first()
    fornecedor = None

    if despesa.fornecedor_id:
        fornecedor = db.query(Fornecedor).filter(Fornecedor.id_fornecedor == despesa.fornecedor_id).first()

    return {
        "edificio": edificio.nome if edificio else "",
        "categoria": categoria.nome_pt if categoria else "",
        "fornecedor": fornecedor.nome if fornecedor else "",
        "valor": f"{float(despesa.valor):.2f} EUR",
        "data": despesa.data_despesa.strftime("%d/%m/%Y") if despesa.data_despesa else "",
    }


def exportar_excel(db: Session,gestor_id: int,data_inicio: Optional[date],data_fim: Optional[date],) -> io.BytesIO:

    validar_gestor(db, gestor_id) # Apenas gestores têm permissão Para exportar dados
    despesas = obter_despesas(db, gestor_id, data_inicio, data_fim)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Despesas"

    cabecalhos = ["Edificio", "Categoria", "Fornecedor", "Valor", "Data"] # colunas
    ws.append(cabecalhos)

    fill = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF") 
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for despesa in despesas:
        ws.append(list(formatar_dados_export(db, despesa).values())) # Preenche as linhas da folha de excel com os dados formatados previamente, tive que fazer isto porque a biblioteca openpyxl não aceita dicionários, tem que ser listas ou tuples

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buffer = io.BytesIO() 
    wb.save(buffer)
    buffer.seek(0) 
    return buffer


def exportar_pdf(db: Session,gestor_id: int,data_inicio: Optional[date],data_fim: Optional[date],) -> io.BytesIO:

    gestor = validar_gestor(db, gestor_id)
    despesas = obter_despesas(db, gestor_id, data_inicio, data_fim, limit=PDF_LIMITE)
    total_bd = (db.query(Despesa).filter(Despesa.status == 1, Despesa.gestor_id == gestor_id).count())
    despesas_total = sum(float(despesa.valor) for despesa in despesas)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    largura, altura = A4
    y = altura - 40

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(40, y, "Sistema de Gestao de Condominios - Relatorio de Despesas")
    y -= 18

    pdf.setFont("Helvetica", 9)
    pdf.drawString(40, y, f"Data de geracao: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    y -= 14
    pdf.drawString(40, y, f"Gestor: {gestor.nome}")
    y -= 14
    pdf.drawString(40, y, f"Registos exportados: {len(despesas)} de {total_bd} disponiveis")
    y -= 14
    pdf.drawString(40, y, f"Montante total: {despesas_total:.2f} EUR")

    if total_bd > PDF_LIMITE:
        pdf.setFillColorRGB(0.8, 0.2, 0.2)
        pdf.drawString(240, y, f"Exportacao limitada a {PDF_LIMITE} registos")
        pdf.setFillColorRGB(0, 0, 0)

    y -= 20
    pdf.setLineWidth(0.5)
    pdf.line(40, y, largura - 40, y)
    y -= 14

    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(40, y, "Edificio")
    pdf.drawString(160, y, "Categoria")
    pdf.drawString(270, y, "Fornecedor")
    pdf.drawString(420, y, "Valor")
    pdf.drawString(490, y, "Data")
    y -= 4
    pdf.line(40, y, largura - 40, y)
    y -= 12

    pdf.setFont("Helvetica", 8)

    for despesa in despesas:
        if y < 60:
            pdf.showPage()
            y = altura - 40
            pdf.setFont("Helvetica", 8)

        linha = formatar_dados_export(db, despesa)
        pdf.drawString(40, y, str(linha["edificio"])[:22])
        pdf.drawString(160, y, str(linha["categoria"])[:20])
        pdf.drawString(270, y, str(linha["fornecedor"])[:26])
        pdf.drawString(420, y, str(linha["valor"]))
        pdf.drawString(490, y, str(linha["data"]))
        y -= 14

    pdf.save()
    buffer.seek(0)
    return buffer
