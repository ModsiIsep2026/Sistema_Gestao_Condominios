import io
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session, joinedload

from app.tabelas_bd.edificio import Edificio
from app.tabelas_bd.apartamento import Apartamento
from app.tabelas_bd.pagamento import Pagamento


def _obter_pagamentos(db: Session, id_gestor: int, data_inicio: Optional[date], data_fim: Optional[date]):
    q = (
        db.query(Pagamento)
        .join(Apartamento, Pagamento.id_apartamento == Apartamento.id)
        .join(Edificio, Apartamento.id_edificio == Edificio.id)
        .options(joinedload(Pagamento.apartamento).joinedload(Apartamento.edificio))
        .filter(
            Edificio.id_gestor == id_gestor,
            Edificio.status == 1,
            Apartamento.status == 1,
            Pagamento.status == 1,
        )
    )
    if data_inicio:
        q = q.filter(Pagamento.data_i >= data_inicio)
    if data_fim:
        q = q.filter(Pagamento.data_i <= data_fim)

    pagamentos = q.all()
    apt_map = {p.id_apartamento: p.apartamento for p in pagamentos if p.apartamento}
    ed_map = {p.apartamento.id_edificio: p.apartamento.edificio for p in pagamentos if p.apartamento and p.apartamento.edificio}

    return pagamentos, apt_map, ed_map


def exportar_excel(db: Session, id_gestor: int, data_inicio: Optional[date], data_fim: Optional[date]) -> io.BytesIO:
    pagamentos, apt_map, ed_map = _obter_pagamentos(db, id_gestor, data_inicio, data_fim)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagamentos"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    cabecalhos = ["Edifício", "Fração", "Mês", "Valor (€)", "Estado", "Data Pagamento"]
    ws.append(cabecalhos)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total = 0.0
    for p in pagamentos:
        apt = apt_map.get(p.id_apartamento)
        ed = ed_map.get(apt.id_edificio) if apt else None
        ws.append([
            ed.rua if ed else "-",
            apt.fracao if apt else "-",
            p.mes,
            float(p.valor),
            "Pago" if p.estado == 1 else "Pendente",
            p.data_p.strftime("%Y-%m-%d") if p.data_p else "-",
        ])
        total += float(p.valor)

    ws.append(["", "", "TOTAL", round(total, 2), "", ""])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_pdf(db: Session, id_gestor: int, data_inicio: Optional[date], data_fim: Optional[date]) -> io.BytesIO:
    pagamentos, apt_map, ed_map = _obter_pagamentos(db, id_gestor, data_inicio, data_fim)

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    largura, altura = A4

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(40, altura - 50, "Relatório de Pagamentos")

    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, altura - 72, f"Período: {data_inicio or 'início'} → {data_fim or 'hoje'}")
    pdf.drawString(40, altura - 88, f"Total de registos: {len(pagamentos)}")
    total = sum(float(p.valor) for p in pagamentos)
    pdf.drawString(40, altura - 104, f"Montante total: {total:.2f} EUR")

    y = altura - 132
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(40,  y, "Edifício")
    pdf.drawString(180, y, "Fração")
    pdf.drawString(250, y, "Mês")
    pdf.drawString(320, y, "Valor")
    pdf.drawString(400, y, "Estado")
    y -= 6
    pdf.line(40, y, largura - 40, y)
    y -= 14

    pdf.setFont("Helvetica", 8)
    for p in pagamentos:
        if y < 60:
            pdf.showPage()
            y = altura - 50
            pdf.setFont("Helvetica", 8)

        apt = apt_map.get(p.id_apartamento)
        ed = ed_map.get(apt.id_edificio) if apt else None
        pdf.drawString(40,  y, (ed.rua[:28] if ed else "-"))
        pdf.drawString(180, y, apt.fracao if apt else "-")
        pdf.drawString(250, y, p.mes)
        pdf.drawString(320, y, f"{float(p.valor):.2f} €")
        pdf.drawString(400, y, "Pago" if p.estado == 1 else "Pendente")
        y -= 14

    pdf.save()
    buffer.seek(0)
    return buffer
